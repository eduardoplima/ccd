"""Importação da planilha `monitoramento_desconto_folha.xlsx`.

3 fluxos:

- **Atualizar IdOrgaoNotificado dos legacy** (aba "Monitoramento Geral"): UPDATE
  em FRAPDescontoFolha.Origem='P' onde IdOrgaoNotificado é NULL.
- **Cadastros novos a partir da aba "Desconto em Folha Valores"**: para CPFs
  que ainda não existem em FRAPDescontoFolha (Ativo=1), cria com Origem='M' +
  parcelas a partir das colunas mensais.
- **Caso Nereu** (aba "Monitoramento NEREU"): cadastro consolidado único com 1
  parcela por multa. Coluna "Desconto verificado no FRAP" alimenta
  FRAPMatchDescontoFolha com IsManual=1 ligando documento → IdLancamentoFRAP.
"""

from __future__ import annotations

import difflib
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import Engine, bindparam, text

CPF_NEREU = "13006444434"
NOME_NEREU = "NEREU BATISTA LINHARES"
SIGLA_NEREU_ORGAO = "IPERN"

# A planilha usa abreviações curtas em "Monitoramento Geral" (IPERN, PGE) e
# nomes completos em "Desconto em Folha Valores". Tabela manual de fallback
# para resolver siglas conhecidas — caso o LIKE em vw_Gen_Orgao não bata.
SIGLA_NOME_ORGAO: dict[str, str] = {
    "IPERN": "INSTITUTO DE PREVIDENCIA DOS SERVIDORES",
    "PGE": "PROCURADORIA GERAL DO ESTADO",
}

# Aliases por IdOrgao: siglas históricas do RN que não casam por nome/código
# (mudaram de nome ou estão fora do padrão). Confirmados em Bdc.dbo.vw_Gen_Orgao:
#   ALRN  = Assembleia Legislativa (id=333, cod='PL', sigla='L003')
#   SEEC  = Secretaria de Educação e Cultura → atual SECD (id=513)
#   SEAD  = Secretaria de Administração → atual SEARH (id=512)
ALIASES_ORGAO_RN: dict[str, int] = {
    "ALRN": 333,
    "SEEC": 513,
    "SEAD": 512,
}

# Documento no extrato vem formatado: "Documento 202.501.310.027.938 do extrato..."
# Tenta variações comuns: 15 dígitos com 4 pontos (3-3-3-3-3), com 5 pontos (4-3-3-3-3),
# ou 15-17 dígitos contíguos sem pontuação. Ordem da lista importa — primeira que casa vence.
_DOC_RES: list[re.Pattern[str]] = [
    re.compile(r"(\d{3}\.\d{3}\.\d{3}\.\d{3}\.\d{3})"),
    re.compile(r"(\d{4}\.\d{3}\.\d{3}\.\d{3}\.\d{3})"),
    re.compile(r"(?<!\d)(\d{15,17})(?!\d)"),
]
_OK_RE = re.compile(r"\bOK\b", re.IGNORECASE)


def _extrai_doc(texto: str) -> str | None:
    """Extrai documento normalizado (só dígitos) do texto. None se nenhum padrão casa."""
    for pat in _DOC_RES:
        m = pat.search(texto)
        if m:
            return re.sub(r"\D", "", m.group(1))
    return None


@dataclass
class ImportResult:
    atualizados_orgao: int = 0
    cadastros_criados: int = 0
    matches_criados: int = 0
    erros: list[dict[str, Any]] = field(default_factory=list)

    def resumo(self) -> str:
        return (
            f"orgao_atualizados={self.atualizados_orgao} "
            f"cadastros_criados={self.cadastros_criados} "
            f"matches_criados={self.matches_criados} "
            f"erros={len(self.erros)}"
        )


# ---------------------------------------------------------------------------
# Parsers (openpyxl, lê por índice de coluna pra evitar issues de encoding)
# ---------------------------------------------------------------------------


def _norm_cpf(v: Any) -> str | None:
    if v is None:
        return None
    digits = "".join(c for c in str(v) if c.isdigit())
    if not digits:
        return None
    digits = digits.zfill(11)
    if len(digits) != 11:
        return None
    return digits


def _str_strip(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def parse_monitoramento_geral(path: Path) -> list[dict[str, Any]]:
    """Aba 'Monitoramento Geral': retorna lista de {cpf, nome, orgao_sigla, processo}.

    Colunas relevantes (1-based): 1=PROCESSO, 13=ÓRGÃO, 21=NOME, 22=CPF.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Monitoramento Geral"]
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        cpf = _norm_cpf(cells[21] if len(cells) > 21 else None)  # col 22 (idx 21)
        if not cpf:
            continue
        out.append(
            {
                "cpf": cpf,
                "nome": _str_strip(cells[20]) if len(cells) > 20 else None,
                "orgao_sigla": _str_strip(cells[12]) if len(cells) > 12 else None,
                "processo": _str_strip(cells[0]) if len(cells) > 0 else None,
            }
        )
    wb.close()
    return out


def parse_desconto_em_folha_valores(path: Path) -> list[dict[str, Any]]:
    """Aba 'Desconto em Folha Valores' wide → long.

    Colunas (1-based): 1=PROCESSO, 3=CPF, 6=ÓRGÃO, 12-28=mensais (dec/2024..abr/2026).
    Retorna 1 dict por pessoa: {cpf, nome (None), orgao, processo, parcelas: [{ano, mes, valor}]}.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Desconto em Folha Valores"]
    # Cabeçalho mensal: capturar (col_idx → (ano, mes)) das colunas 12..28
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    meses_por_col: dict[int, tuple[int, int]] = {}
    for idx, h in enumerate(header):
        if isinstance(h, datetime):
            meses_por_col[idx] = (h.year, h.month)
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        cpf = _norm_cpf(cells[2] if len(cells) > 2 else None)  # col 3 (idx 2)
        if not cpf:
            continue
        parcelas: list[dict[str, Any]] = []
        for col_idx, (ano, mes) in meses_por_col.items():
            if col_idx >= len(cells):
                continue
            v = cells[col_idx]
            if v is None or v == "" or v == 0:
                continue
            try:
                valor = float(v)
            except (TypeError, ValueError):
                continue
            if valor <= 0:
                continue
            parcelas.append({"ano": ano, "mes": mes, "valor": valor})
        out.append(
            {
                "cpf": cpf,
                "nome": None,  # aba "Valores" não traz nome; resolver via "Geral"
                "orgao": _str_strip(cells[5]) if len(cells) > 5 else None,
                "processo": _str_strip(cells[0]) if len(cells) > 0 else None,
                "parcelas": parcelas,
            }
        )
    wb.close()
    return out


def parse_monitoramento_nereu(path: Path) -> list[dict[str, Any]]:
    """Aba 'Monitoramento NEREU': 1 dict por multa.

    Colunas (1-based): 1=número processo, 2=ano processo, 3=VALOR IMPLEMENTADO,
    4=DATA IMPLEMENTAÇÃO, 10='Desconto verificado no FRAP' (texto livre com nº doc).
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Monitoramento NEREU"]
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        # processo + ano
        num_proc = cells[0] if len(cells) > 0 else None
        ano_proc = cells[1] if len(cells) > 1 else None
        valor = cells[2] if len(cells) > 2 else None
        dt_imp = _to_date(cells[3] if len(cells) > 3 else None)
        verif_frap = cells[9] if len(cells) > 9 else None
        if num_proc is None or valor is None or dt_imp is None:
            continue
        try:
            valor_f = float(valor)
        except (TypeError, ValueError):
            continue
        if valor_f <= 0:
            continue
        # Extrai documento ou marca "OK" sem doc (a planilha sinaliza pagamento mas
        # não cita o nº de documento — fallback casa por valor+mês depois).
        doc_norm: str | None = None
        ok_sem_doc = False
        if isinstance(verif_frap, str) and verif_frap.strip():
            doc_norm = _extrai_doc(verif_frap)
            if doc_norm is None and _OK_RE.search(verif_frap):
                ok_sem_doc = True
        out.append(
            {
                "numero_processo": str(num_proc),
                "ano_processo": int(ano_proc) if ano_proc is not None else None,
                "valor": valor_f,
                "data_implementacao": dt_imp,
                "ano": dt_imp.year,
                "mes": dt_imp.month,
                "documento": doc_norm,
                "ok_sem_doc": ok_sem_doc,
            }
        )
    wb.close()
    return out


# ---------------------------------------------------------------------------
# Lookup de órgão (cross-DB Bdc)
# ---------------------------------------------------------------------------


def _candidatos_sigla(chave: str) -> list[str]:
    """Variações comuns: sufixo /RN removido, sem barras/espaços, etc."""
    base = chave.strip().upper()
    out = [base]
    # sufixos comuns
    for suf in ("/RN", "-RN", " RN", "RN", "/RN/", "/ RN"):
        if base.endswith(suf):
            out.append(base[: -len(suf)].strip().rstrip("/-"))
    # parte antes da primeira /
    if "/" in base:
        out.append(base.split("/", 1)[0].strip())
    # remove pontuação/espaços (ex: "SEEC/RN" → "SEECRN")
    compact = re.sub(r"[^A-Z0-9]", "", base)
    if compact and compact != base:
        out.append(compact)
    # dedup preservando ordem
    seen: set[str] = set()
    res: list[str] = []
    for c in out:
        if c and c not in seen:
            seen.add(c)
            res.append(c)
    return res


_indice_codigos_cache: list[tuple[int, str, str, str, int | None]] | None = None


def _carregar_indice_codigos(engine: Engine) -> list[tuple[int, str, str, str, int | None]]:
    """Lazy-loaded: (IdOrgao, CodigoOrgao_UPPER, SiglaOrgao_UPPER, NomeOrgao, IdOrgaoSuperior).

    Só órgãos ativos. Cache em escopo de módulo para evitar N queries.
    """
    global _indice_codigos_cache
    if _indice_codigos_cache is not None:
        return _indice_codigos_cache
    sql = text(
        """
        SELECT IdOrgao,
               UPPER(LTRIM(RTRIM(ISNULL(CodigoOrgao, '')))) AS Codigo,
               UPPER(LTRIM(RTRIM(ISNULL(SiglaOrgao, '')))) AS Sigla,
               LTRIM(RTRIM(NomeOrgao))                     AS Nome,
               IdOrgaoSuperior
        FROM Bdc.dbo.vw_Gen_Orgao
        WHERE OrgaoAtivo = 1
          AND (CodigoOrgao IS NOT NULL OR SiglaOrgao IS NOT NULL)
        """
    )
    with engine.connect() as conn:
        out = [
            (
                int(r[0]),
                str(r[1] or ""),
                str(r[2] or ""),
                str(r[3] or ""),
                int(r[4]) if r[4] is not None else None,
            )
            for r in conn.execute(sql).fetchall()
        ]
    _indice_codigos_cache = out
    return out


def _fuzzy_codigo(
    chave: str,
    indice: list[tuple[int, str, str, str, int | None]],
    *,
    apenas_estaduais: bool = True,
) -> tuple[int, str] | None:
    """Casa `chave` (sigla curta) contra `CodigoOrgao`/`SiglaOrgao` via similaridade.

    Heurísticas para reduzir falsos positivos:
      - exige 4–8 chars na chave (siglas curtas demais ou longas demais não fazem sentido aqui)
      - quando `apenas_estaduais=True`: só considera órgãos com IdOrgaoSuperior=272
      - prioriza códigos com mesma length da chave (transposição/swap de letras)
      - exige ratio ≥ 0.75 e top-1 claramente melhor que top-2 (gap ≥ 0.05)
    """
    n = len(chave)
    if n < 4 or n > 8:
        return None

    # filtra pool
    pool: list[tuple[int, str, str]] = []  # (id, codigo_ou_sigla, nome)
    for id_org, cod, sigla, nome, sup in indice:
        if apenas_estaduais and sup != 272:
            continue
        if cod and cod not in {p[1] for p in pool}:
            pool.append((id_org, cod, nome))
        if sigla and sigla != cod and sigla not in {p[1] for p in pool}:
            pool.append((id_org, sigla, nome))
    if not pool:
        return None

    # ranking: ratio ajustado por penalty de diferença de length
    def _score(c: str) -> float:
        ratio = difflib.SequenceMatcher(None, chave, c).ratio()
        penalty = 0.05 * abs(len(c) - n)  # penaliza length distinto
        return ratio - penalty

    pontuados = sorted(((_score(c), id_, c, nome) for id_, c, nome in pool), reverse=True)
    if not pontuados:
        return None
    top = pontuados[0]
    if top[0] < 0.75:
        return None
    if len(pontuados) >= 2:
        gap = top[0] - pontuados[1][0]
        if gap < 0.05:
            return None
    return top[1], top[3]


def resolver_orgao_por_nome(engine: Engine, nome_ou_sigla: str) -> tuple[int, str] | None:
    """Resolve (IdOrgao, NomeOrgao) em Bdc.dbo.vw_Gen_Orgao.

    Ordem de tentativa:
      1. Match exato em `SiglaOrgao` ou `CodigoOrgao` (com variações `/RN`).
      2. LIKE em `NomeOrgao` com a string original / abreviações expandidas.
      3. LIKE em `NomeOrgao` com nome completo via fallback `SIGLA_NOME_ORGAO`.
      4. Fuzzy (difflib) em `CodigoOrgao`/`SiglaOrgao` — cobre variações como
         `SEEC` → `SECD`, último recurso com cutoff alto.
    """
    if not nome_ou_sigla:
        return None
    chave = nome_ou_sigla.strip().upper()

    # 0) aliases hardcoded (siglas históricas do RN)
    sql_by_id = text(
        "SELECT IdOrgao, LTRIM(RTRIM(NomeOrgao)) FROM Bdc.dbo.vw_Gen_Orgao WHERE IdOrgao = :id"
    )
    with engine.connect() as conn:
        for variante in _candidatos_sigla(chave):
            if variante in ALIASES_ORGAO_RN:
                row = conn.execute(sql_by_id, {"id": ALIASES_ORGAO_RN[variante]}).fetchone()
                if row:
                    return int(row[0]), str(row[1])

    # 1) tenta sigla/código exato primeiro (mais confiável que LIKE)
    sql_sigla = text(
        """
        SELECT TOP 1 IdOrgao, LTRIM(RTRIM(NomeOrgao)) AS NomeOrgao
        FROM Bdc.dbo.vw_Gen_Orgao
        WHERE UPPER(LTRIM(RTRIM(SiglaOrgao))) = :sigla
           OR UPPER(LTRIM(RTRIM(CodigoOrgao))) = :sigla
        ORDER BY OrgaoAtivo DESC, LEN(NomeOrgao)
        """
    )
    with engine.connect() as conn:
        for variante in _candidatos_sigla(chave):
            row = conn.execute(sql_sigla, {"sigla": variante}).fetchone()
            if row:
                return int(row[0]), str(row[1])
    # 2) LIKE no nome — normaliza abreviações comuns
    chave_nome = (
        chave.replace("PREF.", "PREFEITURA")
        .replace("MUN.", "MUNICIPAL")
        .replace("INST.", "INSTITUTO")
        .replace("PREV.", "PREVIDENCIA")
        .replace("SOC.", "SOCIAL")
    )
    # remove sufixo "/RN" se existir
    chave_nome = re.sub(r"/RN\b", "", chave_nome).strip()

    # variação "PREFEITURA DE X" → "PREFEITURA MUNICIPAL DE X"
    chave_nome2: str | None = None
    m = re.match(r"^PREFEITURA\s+DE\s+(.+)$", chave_nome)
    if m:
        chave_nome2 = f"PREFEITURA MUNICIPAL DE {m.group(1)}"
    candidatos: list[str] = [chave_nome]
    if chave_nome2:
        candidatos.append(chave_nome2)
    candidatos.append(chave)
    if chave in SIGLA_NOME_ORGAO:
        candidatos.insert(0, SIGLA_NOME_ORGAO[chave])

    sql = text(
        """
        SELECT TOP 5 IdOrgao, LTRIM(RTRIM(NomeOrgao)) AS NomeOrgao
        FROM Bdc.dbo.vw_Gen_Orgao
        WHERE UPPER(NomeOrgao) LIKE :q
        ORDER BY LEN(NomeOrgao)
        """
    )
    with engine.connect() as conn:
        for c in candidatos:
            rows = conn.execute(sql, {"q": f"%{c}%"}).fetchall()
            if len(rows) == 1:
                return int(rows[0][0]), str(rows[0][1])
            # se a sigla bate exatamente como prefixo, prefere
            if rows:
                for r in rows:
                    nome = str(r[1]).upper()
                    if nome.startswith(c) or f" {c} " in f" {nome} " or f" {c}-" in nome:
                        return int(r[0]), str(r[1])
                # fallback: pega o primeiro (menor nome)
                return int(rows[0][0]), str(rows[0][1])

    # 4) Fuzzy em CodigoOrgao/SiglaOrgao (cobre SEEC→SECD, etc)
    # tenta pela 1ª variante "limpa" da sigla (sem /RN)
    for variante in _candidatos_sigla(chave):
        res = _fuzzy_codigo(variante, _carregar_indice_codigos(engine))
        if res:
            return res
    return None


# ---------------------------------------------------------------------------
# Aplicação no banco
# ---------------------------------------------------------------------------


_SQL_UPDATE_ORGAO_LEGACY = """
UPDATE dbo.FRAPDescontoFolha
SET IdOrgaoNotificado = :id_orgao,
    NomeOrgaoNotificado = :nome_orgao
WHERE CpfCnpj = :cpf
  AND IdOrgaoNotificado IS NULL
  AND Ativo = 1;
"""


def aplicar_orgao_legacy(
    engine: Engine, registros: list[dict[str, Any]], dry_run: bool = False
) -> tuple[int, list[dict[str, Any]]]:
    """Atualiza IdOrgaoNotificado/NomeOrgaoNotificado a partir da planilha.

    Cobre legacy de qualquer Origem ('P' processo, 'C' CCD, 'S' SIAI) que esteja
    com IdOrgaoNotificado NULL. Origem='M' não aparece aqui porque cadastros
    manuais já vêm com órgão obrigatório.

    Retorna (qtd_atualizados, erros).
    """
    erros: list[dict[str, Any]] = []
    cache_orgao: dict[str, tuple[int, str] | None] = {}
    atualizados = 0

    # Quais CPFs ainda têm cadastro sem órgão (qualquer Origem)?
    sql_pendentes = text(
        """
        SELECT DISTINCT CpfCnpj
        FROM dbo.FRAPDescontoFolha
        WHERE Ativo = 1 AND IdOrgaoNotificado IS NULL
          AND CpfCnpj IS NOT NULL
        """
    )
    with engine.connect() as conn:
        pendentes = {str(r[0]) for r in conn.execute(sql_pendentes).fetchall()}

    if not pendentes:
        return 0, erros

    # Mapeia CPF → sigla a partir da planilha (último ganha — assumindo ordem
    # cronológica na aba)
    cpf_para_sigla: dict[str, str] = {}
    for r in registros:
        if r["cpf"] in pendentes and r.get("orgao_sigla"):
            cpf_para_sigla[r["cpf"]] = r["orgao_sigla"]

    if dry_run:
        return len(cpf_para_sigla), erros

    with engine.begin() as conn:
        for cpf, sigla in cpf_para_sigla.items():
            if sigla not in cache_orgao:
                cache_orgao[sigla] = resolver_orgao_por_nome(engine, sigla)
            resolvido = cache_orgao[sigla]
            if resolvido is None:
                erros.append({"cpf": cpf, "motivo": f"orgao não resolvido: {sigla}"})
                continue
            id_org, nome_org = resolvido
            res = conn.execute(
                text(_SQL_UPDATE_ORGAO_LEGACY),
                {"id_orgao": id_org, "nome_orgao": nome_org, "cpf": cpf},
            )
            atualizados += res.rowcount or 0
    return atualizados, erros


_SQL_INSERT_DESCONTO_M = """
INSERT INTO dbo.FRAPDescontoFolha
    (IdDescontoFolha, Origem, CpfCnpj, NomePessoa,
     IdOrgaoNotificado, NomeOrgaoNotificado,
     QtdParcelasPlanejadas, ValorTotalEsperado, Ativo, DataInclusao)
OUTPUT inserted.IdFRAPDescontoFolha
VALUES (NULL, 'M', :cpf, :nome,
        :id_orgao, :nome_orgao,
        :qtd, :valor_total, 1, SYSUTCDATETIME());
"""

_SQL_INSERT_PARCELA_M = """
INSERT INTO dbo.FRAPDescontoFolhaParcela
    (IdFRAPDescontoFolha, IdParcela, NumeroParcela, MesReferencia, AnoReferencia,
     ValorEsperado, DataVencimento)
OUTPUT inserted.IdFRAPDescontoFolhaParcela
VALUES
    (:id_pai, NULL, :numero_parcela, :mes, :ano, :valor, :dt_venc);
"""


def criar_cadastros_novos(
    engine: Engine,
    registros: list[dict[str, Any]],
    nomes_por_cpf: dict[str, str | None],
    dry_run: bool = False,
) -> tuple[int, list[dict[str, Any]]]:
    """Para CPFs não-existentes em FRAPDescontoFolha (Ativo=1), cria com Origem='M'.

    `registros` vem de `parse_desconto_em_folha_valores`.
    `nomes_por_cpf` traz o nome resolvido via aba 'Monitoramento Geral'.
    """
    erros: list[dict[str, Any]] = []
    if not registros:
        return 0, erros

    cpfs = list({r["cpf"] for r in registros if r.get("parcelas")})
    if not cpfs:
        return 0, erros

    with engine.connect() as conn:
        existentes = {
            str(r[0])
            for r in conn.execute(
                text(
                    "SELECT DISTINCT CpfCnpj FROM dbo.FRAPDescontoFolha "
                    "WHERE Ativo = 1 AND CpfCnpj IN :cpfs"
                ).bindparams(bindparam("cpfs", expanding=True)),
                {"cpfs": cpfs},
            ).fetchall()
        }

    novos = [r for r in registros if r["cpf"] not in existentes and r.get("parcelas")]
    if dry_run:
        return len(novos), erros

    cache_orgao: dict[str, tuple[int, str] | None] = {}
    criados = 0
    with engine.begin() as conn:
        for r in novos:
            nome = nomes_por_cpf.get(r["cpf"]) or f"CPF {r['cpf']}"
            orgao_str = r.get("orgao") or ""
            if orgao_str not in cache_orgao:
                cache_orgao[orgao_str] = (
                    resolver_orgao_por_nome(engine, orgao_str) if orgao_str else None
                )
            resolvido = cache_orgao[orgao_str]
            if resolvido is None:
                erros.append({"cpf": r["cpf"], "motivo": f"orgao não resolvido: {orgao_str}"})
                continue
            id_org, nome_org = resolvido
            valor_total = sum(p["valor"] for p in r["parcelas"])
            id_pai = conn.execute(
                text(_SQL_INSERT_DESCONTO_M),
                {
                    "cpf": r["cpf"],
                    "nome": nome,
                    "id_orgao": id_org,
                    "nome_orgao": nome_org,
                    "qtd": len(r["parcelas"]),
                    "valor_total": valor_total,
                },
            ).scalar_one()
            params = [
                {
                    "id_pai": int(id_pai),
                    "numero_parcela": idx + 1,
                    "mes": p["mes"],
                    "ano": p["ano"],
                    "valor": p["valor"],
                    "dt_venc": date(p["ano"], p["mes"], 1),
                }
                for idx, p in enumerate(r["parcelas"])
            ]
            conn.execute(text(_SQL_INSERT_PARCELA_M), params)
            criados += 1
    return criados, erros


# ---------------------------------------------------------------------------
# Caso Nereu — cadastro consolidado + pré-match por documento
# ---------------------------------------------------------------------------


_SQL_FRAPDF_BY_CPF = """
SELECT TOP 1 IdFRAPDescontoFolha
FROM dbo.FRAPDescontoFolha
WHERE CpfCnpj = :cpf AND Ativo = 1 AND Origem = 'M'
ORDER BY IdFRAPDescontoFolha DESC;
"""


def importar_nereu(
    engine: Engine,
    registros: list[dict[str, Any]],
    id_usuario: int,
    dry_run: bool = False,
) -> tuple[int, int, list[dict[str, Any]]]:
    """Cria/atualiza cadastro consolidado do Nereu + pré-matches por documento.

    Retorna (cadastros_criados, matches_criados, erros).
    """
    erros: list[dict[str, Any]] = []
    if not registros:
        return 0, 0, erros

    if dry_run:
        n_doc = sum(1 for r in registros if r.get("documento"))
        return 1, n_doc, erros

    # Resolve órgão IPERN
    orgao = resolver_orgao_por_nome(engine, SIGLA_NEREU_ORGAO)
    if orgao is None:
        erros.append({"cpf": CPF_NEREU, "motivo": "orgao IPERN não resolvido"})
        return 0, 0, erros
    id_org, nome_org = orgao

    cadastros_criados = 0
    matches_criados = 0

    with engine.begin() as conn:
        # Verifica se já existe cadastro
        row = conn.execute(text(_SQL_FRAPDF_BY_CPF), {"cpf": CPF_NEREU}).first()
        if row is None:
            valor_total = sum(r["valor"] for r in registros)
            id_pai = conn.execute(
                text(_SQL_INSERT_DESCONTO_M),
                {
                    "cpf": CPF_NEREU,
                    "nome": NOME_NEREU,
                    "id_orgao": id_org,
                    "nome_orgao": nome_org,
                    "qtd": len(registros),
                    "valor_total": valor_total,
                },
            ).scalar_one()
            cadastros_criados = 1
            # Insere parcelas e captura ids para pré-match
            registros_ord = sorted(
                registros, key=lambda x: (x["ano"], x["mes"], x["numero_processo"])
            )
            id_parcela_por_idx: list[int] = []
            for idx, r in enumerate(registros_ord):
                pid = conn.execute(
                    text(_SQL_INSERT_PARCELA_M),
                    {
                        "id_pai": int(id_pai),
                        "numero_parcela": idx + 1,
                        "mes": r["mes"],
                        "ano": r["ano"],
                        "valor": r["valor"],
                        "dt_venc": date(r["ano"], r["mes"], 1),
                    },
                ).scalar_one()
                id_parcela_por_idx.append(int(pid))
                r["_id_parcela"] = int(pid)
        else:
            # Cadastro já existe — mapeia parcelas existentes e INSERE as faltantes.
            # Sem isso, cadastros antigos sem parcelas (que apareciam como "virtual"
            # na UI antiga) ficam invisíveis em "Por pessoa".
            id_pai = int(row[0])
            existentes = conn.execute(
                text(
                    """
                    SELECT IdFRAPDescontoFolhaParcela, MesReferencia, AnoReferencia,
                           ValorEsperado, NumeroParcela
                    FROM dbo.FRAPDescontoFolhaParcela
                    WHERE IdFRAPDescontoFolha = :id_pai
                    """
                ),
                {"id_pai": id_pai},
            ).fetchall()
            mapa = {(int(e[1]), int(e[2]), float(e[3])): int(e[0]) for e in existentes}
            proximo_numero = max((int(e[4] or 0) for e in existentes), default=0) + 1
            registros_ord = sorted(
                registros, key=lambda x: (x["ano"], x["mes"], x["numero_processo"])
            )
            novas = 0
            for r in registros_ord:
                key = (r["mes"], r["ano"], r["valor"])
                if key in mapa:
                    r["_id_parcela"] = mapa[key]
                    continue
                pid = conn.execute(
                    text(_SQL_INSERT_PARCELA_M),
                    {
                        "id_pai": int(id_pai),
                        "numero_parcela": proximo_numero,
                        "mes": r["mes"],
                        "ano": r["ano"],
                        "valor": r["valor"],
                        "dt_venc": date(r["ano"], r["mes"], 1),
                    },
                ).scalar_one()
                r["_id_parcela"] = int(pid)
                proximo_numero += 1
                novas += 1
            if novas:
                # Atualiza header com total atual de parcelas e valor agregado.
                conn.execute(
                    text(
                        """
                        UPDATE dbo.FRAPDescontoFolha
                        SET QtdParcelasPlanejadas = (
                                SELECT COUNT(*) FROM dbo.FRAPDescontoFolhaParcela
                                WHERE IdFRAPDescontoFolha = :id_pai
                            ),
                            ValorTotalEsperado = (
                                SELECT SUM(ValorEsperado) FROM dbo.FRAPDescontoFolhaParcela
                                WHERE IdFRAPDescontoFolha = :id_pai
                            )
                        WHERE IdFRAPDescontoFolha = :id_pai
                        """
                    ),
                    {"id_pai": id_pai},
                )

        # Pré-matches por documento (ou fallback por valor+mês quando a planilha
        # disse "OK" mas não citou o número do documento).
        id_status_manual = int(
            conn.execute(
                text(
                    "SELECT IdStatusMatch FROM dbo.FRAPStatusMatch "
                    "WHERE Matcher = 'DESCONTO_FOLHA' AND Codigo = 'MATCH_MANUAL'"
                )
            ).scalar_one()
        )
        for r in registros:
            id_parcela = r.get("_id_parcela")
            if not id_parcela:
                continue
            doc = r.get("documento")
            lanc_id: int | None = None
            obs: str = ""
            if doc:
                # FRAPLancamento.Documento vem com pontos do extrato ("202.501.310.027.938").
                # `doc` aqui já está normalizado pra dígitos puros — normaliza no banco também.
                hit = conn.execute(
                    text(
                        """
                        SELECT TOP 1 IdLancamento FROM dbo.FRAPLancamento
                        WHERE REPLACE(REPLACE(REPLACE(REPLACE(
                              Documento, '.', ''), '-', ''), '/', ''), ' ', '') = :doc
                        ORDER BY DtMovimento ASC
                        """
                    ),
                    {"doc": doc},
                ).scalar_one_or_none()
                if hit is None:
                    erros.append(
                        {"cpf": CPF_NEREU, "motivo": f"lancamento não encontrado: doc={doc}"}
                    )
                    continue
                lanc_id = int(hit)
                obs = f"import planilha (doc {doc})"
            elif r.get("ok_sem_doc"):
                # Fallback: planilha sinalizou OK mas sem documento. Casa pelo CPF do
                # depositante + valor exato + mês da implementação (janela de 35 dias).
                ini = date(r["ano"], r["mes"], 1)
                fim = ini + timedelta(days=35)
                candidatos = conn.execute(
                    text(
                        """
                        SELECT IdLancamento FROM dbo.FRAPLancamento
                        WHERE RIGHT(ISNULL(CpfCnpjDepositante, ''), 11) = :cpf
                          AND Valor = :valor
                          AND DtMovimento BETWEEN :ini AND :fim
                        """
                    ),
                    {
                        "cpf": CPF_NEREU,
                        "valor": r["valor"],
                        "ini": ini,
                        "fim": fim,
                    },
                ).fetchall()
                if not candidatos:
                    erros.append(
                        {
                            "cpf": CPF_NEREU,
                            "motivo": f"OK sem doc: nenhum lançamento bate valor/mês ({r['mes']:02d}/{r['ano']} R${r['valor']:.2f})",
                        }
                    )
                    continue
                if len(candidatos) > 1:
                    erros.append(
                        {
                            "cpf": CPF_NEREU,
                            "motivo": f"OK sem doc: {len(candidatos)} candidatos ambíguos para {r['mes']:02d}/{r['ano']} R${r['valor']:.2f}",
                        }
                    )
                    continue
                lanc_id = int(candidatos[0][0])
                obs = (
                    f"import planilha (OK sem doc, casado por valor/mês {r['mes']:02d}/{r['ano']})"
                )
            else:
                continue
            # Idempotente: pula se já existe match (manual ou automático) na mesma combinação
            ja_existe = conn.execute(
                text(
                    "SELECT 1 FROM dbo.FRAPMatchDescontoFolha "
                    "WHERE IdFRAPDescontoFolhaParcela = :ip AND IdLancamentoFRAP = :il"
                ),
                {"ip": id_parcela, "il": lanc_id},
            ).first()
            if ja_existe:
                continue
            conn.execute(
                text(
                    """
                    INSERT INTO dbo.FRAPMatchDescontoFolha
                        (IdFRAPDescontoFolhaParcela, IdLancamentoFRAP, IdStatusMatch,
                         IsManual, IdUsuarioConcilia, DataConcilia, Observacao)
                    VALUES (:ip, :il, :istatus, 1, :iu, SYSUTCDATETIME(),
                            :obs)
                    """
                ),
                {
                    "ip": id_parcela,
                    "il": lanc_id,
                    "istatus": id_status_manual,
                    "iu": id_usuario,
                    "obs": obs,
                },
            )
            matches_criados += 1

    return cadastros_criados, matches_criados, erros


# ---------------------------------------------------------------------------
# Orquestração
# ---------------------------------------------------------------------------


def importar_planilha(
    engine: Engine,
    arquivo: Path,
    *,
    apenas_orgao: bool = False,
    criar_cadastros: bool = False,
    incluir_nereu: bool = False,
    id_usuario: int | None = None,
    dry_run: bool = False,
) -> ImportResult:
    """Roda os 3 fluxos conforme flags. Sem flags, roda todos."""
    rodar_tudo = not (apenas_orgao or criar_cadastros or incluir_nereu)
    res = ImportResult()

    geral = parse_monitoramento_geral(arquivo)
    nomes_por_cpf = {r["cpf"]: r["nome"] for r in geral if r.get("nome")}

    if apenas_orgao or rodar_tudo:
        n, errs = aplicar_orgao_legacy(engine, geral, dry_run=dry_run)
        res.atualizados_orgao = n
        res.erros.extend(errs)

    if criar_cadastros or rodar_tudo:
        valores = parse_desconto_em_folha_valores(arquivo)
        n, errs = criar_cadastros_novos(engine, valores, nomes_por_cpf, dry_run=dry_run)
        res.cadastros_criados += n
        res.erros.extend(errs)

    if incluir_nereu or rodar_tudo:
        if id_usuario is None and not dry_run:
            res.erros.append({"cpf": CPF_NEREU, "motivo": "id_usuario obrigatório p/ Nereu"})
        else:
            nereu = parse_monitoramento_nereu(arquivo)
            cads, matches, errs = importar_nereu(engine, nereu, id_usuario or 0, dry_run=dry_run)
            res.cadastros_criados += cads
            res.matches_criados += matches
            res.erros.extend(errs)

    return res
