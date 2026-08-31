"""Export do staging para o lote de importação do SisBenefícios.

Gera xlsx e json com os NOMES DE COLUNA de
BdBeneficio.dbo.Beneficio_PropostaBeneficio, 1:1 para o script importador do
outro setor. Duas colunas extras de correlação: IdInterno (IdCCDBeneficio) e
IdInternoBeneficioAnterior (vínculo efetivo->potencial, que o importador
resolve para IdBeneficioAnterior após inserir os potenciais do lote).
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import bindparam, text
from sqlalchemy.orm import Session

from app.ccd.beneficios.dominios import obter_dominios
from app.config import get_settings

# staging -> nome no Beneficio_PropostaBeneficio (ordem do arquivo exportado)
_COLUNAS_EXPORT: list[tuple[str, str]] = [
    ("IdCCDBeneficio", "IdInterno"),
    ("IdCCDBeneficioPotencial", "IdInternoBeneficioAnterior"),
    ("DescricaoPropostaBeneficio", "DescricaoPropostaBeneficio"),
    ("MemoriaCalculoPropostaBeneficio", "MemoriaCalculoPropostaBeneficio"),
    ("ValorQuantidade", "ValorQuantidade"),
    ("JustificativaPropostaBeneficio", "JustificativaPropostaBeneficio"),
    ("IdBeneficioSituacaoEfetivacao", "IdBeneficioSituacaoEfetivacao"),
    ("IdAreaTematica", "IdAreaTematica"),
    ("IdCaracterizacaoBeneficio", "IdCaracterizacaoBeneficio"),
    ("IdUnidadeDeMedida", "IdUnidadeDeMedida"),
    ("IdBeneficioSituacao", "IdBeneficioSituacao"),
    ("IdTipoBeneficio", "IdTipoBeneficio"),
    ("IdSubTipoBeneficio", "IdSubTipoBeneficio"),
    ("NumeroProcessoDecisao", "NumeroProcessoDecisao"),
    ("AnoProcessoDecisao", "AnoProcessoDecisao"),
    ("IdProcessoDecisao", "IdProcessoDecisao"),
    ("DescricaoMotivo", "DescricaoMotivo"),
]

# IdStatusBeneficio do workflow do SisBenefícios: todo lote entra como 1=Cadastrado.
_ID_STATUS_CADASTRADO = 1

# Identidade visual TCE/RN
_VERDE = "2E5B3C"
_VERDE_ESCURO = "1A3D28"
_CINZA_CLARO = "F2F2F2"


class ExportVazio(Exception):
    pass


def _selecionar(session: Session, ids: list[int] | None) -> list[dict[str, Any]]:
    where = "b.Ativo = 1 AND b.Status = 'VALIDADO'"
    params: dict[str, Any] = {}
    if ids:
        where += " AND b.IdCCDBeneficio IN :ids"
        params["ids"] = ids
    stmt = text(f"SELECT b.* FROM dbo.CCDBeneficio b WHERE {where} ORDER BY b.IdCCDBeneficio")
    if ids:
        stmt = stmt.bindparams(bindparam("ids", expanding=True))
    return [dict(r) for r in session.execute(stmt, params).mappings().all()]


def _linha_export(row: dict[str, Any], id_setor: int | None) -> dict[str, Any]:
    linha = {destino: row[origem] for origem, destino in _COLUNAS_EXPORT}
    linha["IdStatusBeneficio"] = _ID_STATUS_CADASTRADO
    linha["IdSetorUsuarioCadastro"] = id_setor
    # Origem PROPOSTA: o registro deriva de uma proposta real do BdBeneficio —
    # IdBeneficioAnterior recebe o id dela (ChaveOrigem = 'PROPOSTA:<id>') para
    # o importador vincular a conversão proposta -> potencial/efetivo.
    chave = row.get("ChaveOrigem") or ""
    linha["IdBeneficioAnterior"] = (
        int(chave.split(":", 1)[1]) if row.get("Origem") == "PROPOSTA" and ":" in chave else None
    )
    return linha


def _marcar_enviados(session: Session, ids: list[int], lote: str, id_usuario: int) -> None:
    session.execute(
        text(
            """
            UPDATE dbo.CCDBeneficio
            SET Status = 'ENVIADO',
                DataEnvio = SYSUTCDATETIME(),
                LoteEnvio = :lote,
                DataAtualizacao = SYSUTCDATETIME(),
                IdUsuarioAtualizacao = :id_usuario
            WHERE IdCCDBeneficio IN :ids
            """
        ).bindparams(bindparam("ids", expanding=True)),
        {"lote": lote, "id_usuario": id_usuario, "ids": ids},
    )


def _json_default(v: Any) -> Any:
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, datetime):
        return v.isoformat()
    raise TypeError(f"não serializável: {type(v)}")


def _xlsx(
    linhas: list[dict[str, Any]], rows_staging: list[dict[str, Any]], session: Session
) -> bytes:
    dominios = obter_dominios(session)
    desc = {
        "tipo": {d.id: d.descricao for d in dominios.tipos},
        "subtipo": {d.id: d.descricao for d in dominios.subtipos},
        "area": {d.id: d.descricao for d in dominios.areas_tematicas},
        "caract": {d.id: d.descricao for d in dominios.caracterizacoes},
        "situacao": {d.id: d.descricao for d in dominios.situacoes},
        "efetivacao": {d.id: d.descricao for d in dominios.situacoes_efetivacao},
        "unidade": {d.id: d.descricao for d in dominios.unidades_medida},
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Importacao"
    header_fill = PatternFill("solid", fgColor=_VERDE)
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    alt_fill = PatternFill("solid", fgColor=_CINZA_CLARO)

    colunas = [destino for _, destino in _COLUNAS_EXPORT] + [
        "IdBeneficioAnterior",
        "IdStatusBeneficio",
        "IdSetorUsuarioCadastro",
    ]
    ws.append(colunas)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for i, linha in enumerate(linhas):
        ws.append([linha[c] for c in colunas])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = alt_fill
    for j, c in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(14, min(len(c) + 2, 34))

    # Aba de conferência humana: descrições no lugar dos IDs de domínio.
    ws2 = wb.create_sheet("Conferencia")
    cab2 = [
        "IdInterno",
        "Processo",
        "Pessoa",
        "CpfCnpj",
        "Estágio",
        "Tipo",
        "Subtipo",
        "Área temática",
        "Característica",
        "Situação",
        "Unidade",
        "Valor/Qtd",
        "Data ocorrência",
        "Origem",
        "Descrição",
    ]
    ws2.append(cab2)
    for cell in ws2[1]:
        cell.fill = PatternFill("solid", fgColor=_VERDE_ESCURO)
        cell.font = header_font
    for r in rows_staging:
        proc = (
            f"{r['NumeroProcessoDecisao']}/{r['AnoProcessoDecisao']}"
            if r["NumeroProcessoDecisao"]
            else None
        )
        ws2.append(
            [
                r["IdCCDBeneficio"],
                proc,
                r["NomePessoa"],
                r["CpfCnpj"],
                desc["efetivacao"].get(r["IdBeneficioSituacaoEfetivacao"]),
                desc["tipo"].get(r["IdTipoBeneficio"]),
                desc["subtipo"].get(r["IdSubTipoBeneficio"]),
                desc["area"].get(r["IdAreaTematica"]),
                desc["caract"].get(r["IdCaracterizacaoBeneficio"]),
                desc["situacao"].get(r["IdBeneficioSituacao"]),
                desc["unidade"].get(r["IdUnidadeDeMedida"]),
                r["ValorQuantidade"],
                r["DataOcorrencia"],
                r["Origem"],
                r["DescricaoPropostaBeneficio"],
            ]
        )
    for j, c in enumerate(cab2, start=1):
        ws2.column_dimensions[get_column_letter(j)].width = max(12, min(len(c) + 6, 40))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar(
    session: Session,
    *,
    ids: list[int] | None,
    formato: str,
    marcar_enviado: bool,
    id_usuario: int,
) -> tuple[str, bytes, str]:
    """Retorna (nome_arquivo, conteúdo, media_type). Exporta os VALIDADO
    (todos, ou a interseção com `ids`); marca ENVIADO na mesma transação."""
    rows = _selecionar(session, ids)
    if not rows:
        raise ExportVazio
    lote = "beneficios-ccd-" + datetime.now(UTC).strftime("%Y%m%d-%H%M")
    id_setor = get_settings().beneficio_id_setor_ccd
    linhas = [_linha_export(r, id_setor) for r in rows]

    if formato == "json":
        payload = {
            "lote": lote,
            "geradoEm": datetime.now(UTC).isoformat(),
            "origem": "CCD/DIP - staging CCDBeneficio (BdDIP)",
            "itens": linhas,
        }
        conteudo = json.dumps(payload, ensure_ascii=False, indent=2, default=_json_default).encode(
            "utf-8"
        )
        nome, media = f"{lote}.json", "application/json"
    else:
        conteudo = _xlsx(linhas, rows, session)
        nome = f"{lote}.xlsx"
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if marcar_enviado:
        _marcar_enviados(session, [int(r["IdCCDBeneficio"]) for r in rows], lote, id_usuario)
    session.commit()
    return nome, conteudo, media
