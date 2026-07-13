"""Atualiza scripts/analise/docs/analise_debitos_nereu_definitiva.xlsx.

- Débitos JÁ na planilha: recalcula `valor_multa` e `status_divida` do banco (por id_debito).
- Débitos NOVOS (multa do Nereu ausente da planilha): monta a linha completa (banco +
  decisões + servidores/órgãos + regras) e roda o LLM SÓ neles (`trechos_verba_voto`),
  inserindo em TodosDebitos e na aba de detalhe conforme a regra do vantagens_trans.ipynb.
- Preserva coluna `dasa` (OutrosOrgaos), coluna `vt1` e a aba `Resumo` (pivot): edição in
  place via openpyxl (pandas to_excel destruiria o pivot).

Uso:
    python -m scripts.analise.atualizar_debitos_nereu_definitiva --dry-run
    python -m scripts.analise.atualizar_debitos_nereu_definitiva            # gera _atualizada.xlsx
    python -m scripts.analise.atualizar_debitos_nereu_definitiva --inplace  # grava por cima

As queries são copiadas verbatim das cells do notebook; o CPF do Nereu é hardcoded.
"""
from __future__ import annotations

import argparse
import os
from pathlib import Path

import openpyxl
import pandas as pd

from ccd.db import get_connection

# gpt-4o (que o notebook usava) não está deployado neste recurso Azure, e o endpoint
# em scripts/.env é o surface OpenAI-compatível (/openai/v1) — precisa de ChatOpenAI com
# base_url, não AzureChatOpenAI. gpt-4.1 é o deployment disponível equivalente.
DEFAULT_LLM_MODEL = "gpt-4.1"

CPF_NEREU = "13006444434"
DOCS = Path(__file__).resolve().parent / "docs"
XLSX = DOCS / "analise_debitos_nereu_definitiva.xlsx"
DATA_SHEETS = ["TodosDebitos", "SesapVerbaTransitoria", "SesapOutrosAssuntos", "OutrosOrgaos"]
SEM_VERBA = "Não há verba transitória mencionada no voto."

# --- cell 76: débitos do Nereu (valor atualizado, relator, status da dívida) ---------
SQL_DEBITOS = f"""
SELECT DISTINCT
        ed.IdDebito as id_debito,
        CONCAT(pro_exe.numero_processo, '/', pro_exe.ano_processo) as processo_execucao,
        etd.Descricao as tipo_multa,
        processo.dbo.fn_Exe_RetornaValorAtualizado(ed.IdDebito) as valor_multa,
        r.nome as relator,
        (SELECT CONCAT(numero_processo, '/', ano_processo)
         FROM processo.dbo.Processos WHERE IdProcesso = ed.IdProcessoOrigem) as processo_origem,
        pro_orig.codigo_tipo_processo as tipo_processo_origem,
        (SELECT o.nome FROM processo.dbo.Orgaos o
         WHERE o.IdOrgao = pro_orig.IdOrgaoEnvolvido) as orgao_envolvido_processo_origem,
        esd.DescricaoStatusDivida as status_divida
    FROM processo.dbo.Exe_Debito ed
    LEFT JOIN processo.dbo.Processos pro_orig ON ed.IdProcessoOrigem = pro_orig.IdProcesso
    LEFT JOIN processo.dbo.Processos pro_exe ON ed.IdProcessoExecucao = pro_exe.IdProcesso
    LEFT JOIN processo.dbo.Relator r ON r.codigo = pro_orig.codigo_relator
    LEFT JOIN processo.dbo.Exe_DebitoPessoa edp ON edp.IDDebito = ed.IdDebito
    LEFT JOIN processo.dbo.Exe_TipoDebito etd ON etd.CodigoTipoDebito = ed.CodigoTipoDebito
    LEFT JOIN processo.dbo.GenPessoa gp ON gp.IdPessoa = edp.IDPessoa
    LEFT JOIN processo.dbo.Exe_StatusDivida esd ON esd.CodigoStatusDivida = ed.CodigoStatusDivida
    LEFT JOIN processo.dbo.Exe_Debito_MultaCominatoria edmc ON ed.IdDebito = edmc.IdDebito
    WHERE gp.Documento = '{CPF_NEREU}'
"""


def _in_list(values) -> str:
    """', '.join dos processos com aspas — igual ao notebook (valores vêm do banco)."""
    return ", ".join(f"'{v}'" for v in values)


def is_natureza_transitoria(ementa) -> bool:  # cell 93
    if pd.isna(ementa):
        return False
    keywords = ["propter lab", "natureza transit", "vantagens transit", "vantagem transit", "insalubrida"]
    return any(k in ementa.lower() for k in keywords)


def is_sesap(orgaos_servidores, orgao_envolvido) -> str:  # cell 124
    try:
        if "SESAP" in (orgaos_servidores or "") or "SAÚDE" in (orgao_envolvido or ""):
            return "SESAP"
    except Exception:
        pass
    return "OUTROS"


def build_llm(model: str):
    from langchain_openai import ChatOpenAI

    return ChatOpenAI(
        base_url=os.environ["AZURE_OPENAI_ENDPOINT"],
        api_key=os.environ["AZURE_OPENAI_API_KEY"],
        model=model,
        temperature=0.0,
    )


def get_verba_voto(llm, texto: str) -> str:  # cell 96
    from langchain.prompts import PromptTemplate

    prompt = PromptTemplate.from_template(
        """
        Você é um agente que analisa e categoriza votos de decisões do TCE/RN.
        Seu objetivo é decidir se o voto é sobre uma verba transitória ou não.
        Verbas transitórias são aquelas que possuem natureza transitória, como propter
        laborem, vantagens transitórias, insalubridade, etc.
        O texto do voto é o seguinte:
        "{input}"

        Encontre os trechos onde trata de verba transitória. Separe os trechos encontrados
        com uma quebra de linha.
        Se não encontrar, responda "Não há verba transitória mencionada no voto".

        Sua resposta:
        """
    )
    try:
        return (prompt | llm).invoke(texto).content
    except Exception as e:  # não derruba o lote inteiro por causa de um voto
        print(f"  [LLM erro] {e}")
        return SEM_VERBA


def _no_verba(trecho) -> bool:
    return pd.isna(trecho) or "não há verba transitória" in str(trecho).lower()


def load_sheet_ids(wb) -> set[int]:
    ids: set[int] = set()
    for name in DATA_SHEETS:
        ws = wb[name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = header.index("id_debito") + 1
        for row in ws.iter_rows(min_row=2):
            v = row[col - 1].value
            if v is not None:
                ids.add(int(v))
    return ids


def build_new_rows(engine, llm, novos: pd.DataFrame) -> pd.DataFrame:
    """Monta as linhas completas dos débitos novos (com LLM em trechos_verba_voto)."""
    processos = [p for p in novos["processo_origem"].dropna().unique() if p and p != "/"]
    lista = _in_list(processos)

    # decisões (cells 86/88/90)
    dec_cols = ["texto_acordaos", "ementas_decisoes", "relatorios_decisoes", "votos_decisoes"]
    if processos:
        df_dec = pd.read_sql(
            f"""SELECT numeroprocesso, anoprocesso, fundamentacaovoto, ementa, texto_acordao, relatorio
                FROM processo.dbo.vw_ia_votos_acordaos_decisoes
                WHERE concat(numeroprocesso, '/', anoprocesso) in ({lista}) AND VotoEscolhido = 1""",
            engine,
        )
    else:
        df_dec = pd.DataFrame(columns=["numeroprocesso", "anoprocesso", "fundamentacaovoto", "ementa", "texto_acordao", "relatorio"])
    if len(df_dec):
        df_dec["processo"] = df_dec["numeroprocesso"].astype(str) + "/" + df_dec["anoprocesso"].astype(str)
        dec_grp = df_dec.groupby("processo").agg({
            "texto_acordao": lambda x: "\n".join(x.dropna().astype(str)),
            "ementa": lambda x: "\n".join(x.dropna().astype(str)),
            "relatorio": lambda x: "\n".join(x.dropna().astype(str)),
            "fundamentacaovoto": lambda x: "\n".join(x.dropna().astype(str)),
        }).reset_index().rename(columns={
            "texto_acordao": "texto_acordaos", "ementa": "ementas_decisoes",
            "relatorio": "relatorios_decisoes", "fundamentacaovoto": "votos_decisoes",
        })
    else:
        dec_grp = pd.DataFrame(columns=["processo", *dec_cols])

    # servidores/órgãos envolvidos (cells 102-108)
    if processos:
        df_resp = pd.read_sql(
            f"""SELECT CONCAT(p.numero_processo, '/', p.ano_processo) as processo,
                       gp.Documento as cpf, gp.Nome COLLATE SQL_Latin1_General_CP1_CI_AS as nome
                FROM processo.dbo.Processos p
                INNER JOIN processo.dbo.Pro_ProcessosResponsavelDespesa pprd ON p.IdProcesso = pprd.IdProcesso
                INNER JOIN processo.dbo.GenPessoa gp ON gp.IdPessoa = pprd.IdPessoa
                WHERE CONCAT(p.numero_processo, '/', p.ano_processo) IN ({lista})
                AND gp.Documento NOT IN ('08242034000102', '{CPF_NEREU}') AND len(gp.Documento) = 11""",
            engine,
        )
    else:
        df_resp = pd.DataFrame(columns=["processo", "cpf", "nome"])

    cpfs = df_resp["cpf"].unique()
    if len(cpfs):
        df_pess = pd.read_sql(
            f"""SELECT DISTINCT codigo_orgao, cpf FROM BdDIP.dbo.vwSiaiPessoalFolhaResumida fr
                WHERE fr.CPF IN ({_in_list(cpfs)}) AND fr.codigo_orgao NOT IN ('IPERN')""",
            engine,
        )
    else:
        df_pess = pd.DataFrame(columns=["codigo_orgao", "cpf"])

    if len(df_resp):
        pess_grp = df_pess.groupby("cpf").agg({"codigo_orgao": pd.Series.tolist}).reset_index()
        orgaos = df_resp.merge(pess_grp, how="left", on="cpf")[["processo", "nome", "codigo_orgao"]]
        orgaos = orgaos.dropna(subset=["codigo_orgao"])
        orgaos = orgaos.groupby("processo").agg(lambda x: list(x)).reset_index()
        orgaos["orgaos_servidores_envolvidos"] = orgaos["codigo_orgao"].apply(
            lambda lst: ", ".join(o for sub in lst for o in sub) if isinstance(lst, list) else lst
        )
        orgaos["servidores_envolvidos"] = orgaos["nome"].apply(
            lambda x: ", ".join(x) if isinstance(x, list) else x
        )
        orgaos = orgaos[["processo", "servidores_envolvidos", "orgaos_servidores_envolvidos"]]
    else:
        orgaos = pd.DataFrame(columns=["processo", "servidores_envolvidos", "orgaos_servidores_envolvidos"])

    df = novos.merge(dec_grp, how="left", left_on="processo_origem", right_on="processo").drop(columns=["processo"], errors="ignore")
    df = df.merge(orgaos, how="left", left_on="processo_origem", right_on="processo").drop(columns=["processo"], errors="ignore")

    df["servidores_envolvidos"] = df["servidores_envolvidos"].fillna("Sem servidores envolvidos")
    df["orgaos_servidores_envolvidos"] = df["orgaos_servidores_envolvidos"].fillna("Sem órgãos envolvidos")
    for c in dec_cols:
        df[c] = df[c].fillna("Sem decisões")
    df["verbas_transitorias"] = df["ementas_decisoes"].apply(
        lambda e: "Verba Transitória" if is_natureza_transitoria(e) else "Outros assuntos"
    )
    df.loc[df["ementas_decisoes"] == "Sem decisões", "verbas_transitorias"] = "Inconclusivo"
    df["sesap"] = df.apply(lambda r: is_sesap(r["orgaos_servidores_envolvidos"], r["orgao_envolvido_processo_origem"]), axis=1)

    print(f"  Rodando LLM em {len(df)} débitos novos...")
    df["trechos_verba_voto"] = df["votos_decisoes"].apply(lambda t: get_verba_voto(llm, t))
    # vt1 (curadoria fina) não é derivável do regex verbas_transitorias — usa o mesmo
    # sinal do LLM que o notebook usa p/ separar abas. Auto-preenchido, revisar depois.
    def _vt1(r):
        if r["verbas_transitorias"] == "Inconclusivo":
            return "Inconclusivo"
        return "Outros assuntos" if _no_verba(r["trechos_verba_voto"]) else "Verba transitória"
    df["vt1"] = df.apply(_vt1, axis=1)
    df["dasa"] = None
    return df


def build_summary_sheet(wb) -> None:
    """(Re)cria a aba ResumoDados calculada direto de TodosDebitos, com separação por
    estado (status_divida) × sesap × vt1. Substitui a dependência do pivot Resumo."""
    from collections import defaultdict

    ws = wb["TodosDebitos"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {h: i for i, h in enumerate(header)}
    agg: dict = defaultdict(lambda: [0, 0.0])
    for r in ws.iter_rows(min_row=2, values_only=True):
        key = (r[ci["status_divida"]], r[ci["sesap"]], r[ci["vt1"]])
        raw = r[ci["valor_multa"]]
        try:
            val = float(raw) if pd.notna(raw) else 0.0
        except (TypeError, ValueError):
            val = 0.0
        a = agg[key]
        a[0] += 1
        a[1] += val

    if "ResumoDados" in wb.sheetnames:
        del wb["ResumoDados"]
    out = wb.create_sheet("ResumoDados")
    out.append(["status_divida", "sesap", "vt1", "n_debitos", "valor_total"])
    cur = None
    sub = [0, 0.0]
    por_estado: dict = {}
    for key in sorted(agg, key=lambda k: tuple(str(x) for x in k)):
        st, se, v1 = key
        n, val = agg[key]
        if cur is not None and st != cur:
            out.append([f"Total {cur}", "", "", sub[0], sub[1]])
            por_estado[cur] = tuple(sub)
            sub = [0, 0.0]
        cur = st
        out.append([st, se, v1, n, val])
        sub[0] += n
        sub[1] += val
    if cur is not None:
        out.append([f"Total {cur}", "", "", sub[0], sub[1]])
        por_estado[cur] = tuple(sub)

    # Totais claros: o bruto soma cancelados/pagos (não é dívida) — separar.
    def soma(pred):
        ns = [v[0] for k, v in por_estado.items() if pred(k)]
        vs = [v[1] for k, v in por_estado.items() if pred(k)]
        return sum(ns), sum(vs)
    ativo = lambda k: not (str(k).startswith("Cancel") or str(k).startswith("Pago"))
    out.append([None, None, None, None, None])
    out.append(["Total EM ABERTO (dívida atual)", "", "", *soma(lambda k: k == "Em Aberto")])
    out.append(["Total Em Aberto + Suspenso", "", "", *soma(lambda k: k in ("Em Aberto", "Suspenso"))])
    out.append(["Total exceto cancelados/pagos", "", "", *soma(ativo)])
    out.append(["Total BRUTO (todos os estados)", "", "", *soma(lambda k: True)])


def target_detail_sheet(row) -> str | None:
    """Regra de colocação do notebook (cells 135/136)."""
    if "Aberto" not in str(row["status_divida"]):
        return None
    if row["sesap"] == "SESAP":
        if row["verbas_transitorias"] == "Outros assuntos":
            return "SesapOutrosAssuntos" if _no_verba(row["trechos_verba_voto"]) else "SesapVerbaTransitoria"
        return None  # SESAP + Verba Transitória: só TodosDebitos
    return "OutrosOrgaos"


def append_row(ws, header, rowdict) -> None:
    # NaN (valor NULL no banco) → None; openpyxl não serializa NaN corretamente.
    ws.append([None if pd.isna(v := rowdict.get(h)) else v for h in header])


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default=str(XLSX))
    ap.add_argument("--out", default=None, help="destino (default: <xlsx>_atualizada.xlsx)")
    ap.add_argument("--inplace", action="store_true", help="grava por cima do original")
    ap.add_argument("--dry-run", action="store_true", help="não grava e não chama o LLM")
    ap.add_argument("--model", default=DEFAULT_LLM_MODEL, help="deployment LLM (default: gpt-4.1)")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    out = xlsx if args.inplace else Path(args.out) if args.out else xlsx.with_name(xlsx.stem + "_atualizada.xlsx")

    engine = get_connection()
    llm = None if args.dry_run else build_llm(args.model)

    df_deb = pd.read_sql(SQL_DEBITOS, engine).drop_duplicates("id_debito")
    db_map = df_deb.set_index("id_debito")[["valor_multa", "status_divida"]].to_dict("index")
    print(f"Banco: {len(df_deb)} débitos do Nereu.")

    wb = openpyxl.load_workbook(xlsx)
    ids_planilha = load_sheet_ids(wb)
    assert any(int(i) in ids_planilha for i in df_deb["id_debito"]), "Nenhum id_debito casou — join/coluna errada?"

    # --- antigos: recalcula valor_multa + status_divida ---
    atualizados = {}
    for name in DATA_SHEETS:
        ws = wb[name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        cid, cval, cst = (header.index("id_debito"), header.index("valor_multa"), header.index("status_divida"))
        n = 0
        for row in ws.iter_rows(min_row=2):
            idv = row[cid].value
            if idv is None or int(idv) not in db_map:
                continue
            new = db_map[int(idv)]
            row[cval].value = float(new["valor_multa"]) if pd.notna(new["valor_multa"]) else None
            row[cst].value = new["status_divida"]
            n += 1
        atualizados[name] = n

    orfaos = sorted(ids_planilha - set(int(i) for i in df_deb["id_debito"]))
    novos_df = df_deb[~df_deb["id_debito"].isin(ids_planilha)].copy()

    print("\n== Antigos recalculados por aba ==")
    for name, n in atualizados.items():
        print(f"  {name}: {n}")
    print(f"\nÓrfãos (na planilha, fora do banco): {len(orfaos)} {orfaos[:20]}")
    print(f"Novos débitos a inserir: {len(novos_df)}")

    if len(novos_df) == 0:
        colocacao = {}
    else:
        if args.dry_run:
            print("  (dry-run: pulando LLM — colocação usa só status/sesap parcial)")
            novos_full = None
            colocacao = {"(dry-run, LLM pulado)": len(novos_df)}
        else:
            novos_full = build_new_rows(engine, llm, novos_df)
            colocacao = {s: 0 for s in DATA_SHEETS}
            headers = {name: [c.value for c in next(wb[name].iter_rows(min_row=1, max_row=1))] for name in DATA_SHEETS}
            for _, r in novos_full.iterrows():
                d = r.to_dict()
                append_row(wb["TodosDebitos"], headers["TodosDebitos"], d)
                colocacao["TodosDebitos"] += 1
                dest = target_detail_sheet(r)
                if dest:
                    append_row(wb[dest], headers[dest], d)
                    colocacao[dest] += 1
            print("\n== Novos inseridos por aba ==")
            for name in DATA_SHEETS:
                print(f"  {name}: {colocacao[name]}")
            print("Novos id_debito:", list(novos_full["id_debito"]))

    if args.dry_run:
        print("\n[dry-run] nada foi gravado.")
        return

    build_summary_sheet(wb)
    print("\nAba 'ResumoDados' recalculada dos dados (separação por estado).")

    wb.save(out)
    print(f"\nGravado: {out}")
    if not args.inplace:
        print("Confira a cópia, atualize o pivot 'Resumo' no Excel, preencha vt1/dasa dos novos e renomeie por cima do original.")


if __name__ == "__main__":
    main()
